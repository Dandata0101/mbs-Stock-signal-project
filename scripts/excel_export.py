import os
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

def filter_date_range(df):
    """Filter the DataFrame to include data from Jan 1, 2020, to the present, retaining the 'Date' column."""
    start_date = pd.to_datetime('2020-01-01').date()
    if 'Date' in df.columns:
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce').dt.date
        mask = df['Date'] >= start_date
        filtered_df = df[mask].copy()
    else:
        raise ValueError("DataFrame must have a 'Date' column.")
    return filtered_df


def autofit_columns_for_all_sheets(workbook):
    """Adjust column widths to fit content for all sheets in the workbook."""
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        column_widths = {}
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value:
                    # Calculate the max length of content in each column
                    column_widths[cell.column_letter] = max(
                        column_widths.get(cell.column_letter, 0), 
                        len(str(cell.value))
                    )
        for column, width in column_widths.items():
            adjusted_width = (width + 2) * 1.1  # Adjust the width slightly to avoid cutting off text
            worksheet.column_dimensions[column].width = adjusted_width

def export_df_to_excel_with_chart(df=None, tickerSymbol=None, output_directory='03-output', 
                                  chart_path='static/images/chart.png', stats_csv_path='model_stats.csv'):
    print('4.1) Prepping chart and excel export')

    df = filter_date_range(df)
    current_directory = os.getcwd()
    directory_path = os.path.join(current_directory, output_directory)
    pattern = os.path.join(directory_path, '*stock*.xlsx')

    excel_files = glob.glob(pattern)
    for file_path in excel_files:
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f'4.2 Deleted: {file_path}')

    excel_path = os.path.join(directory_path, f'{tickerSymbol}_stock.xlsx')
    stats_df = pd.read_excel(os.path.join(current_directory, '01-data', 'accuracy_export.xlsx'))
    features_df = pd.read_excel(os.path.join(current_directory, '01-data', 'feature_export.xlsx'))

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Data', startrow=28, startcol=1,index=False)
        stats_df.to_excel(writer, sheet_name='model_stats', startrow=4, startcol=2,index=False)
        features_df.to_excel(writer, sheet_name='model_stats', startrow=10, startcol=2,index=False)

    wb = load_workbook(excel_path)
    ws = wb['Data']

    img = Image(chart_path)
    ws.add_image(img, 'B2')
    ws.sheet_view.showGridLines = False

    # Ensure gridlines are removed from the 'model_stats' sheet as well
    ws_stats = wb['model_stats']
    stats_img = Image(current_directory+'/static/images/confusion_matrix.png')  # Ensure this path is correct
    ws_stats.add_image(stats_img, 'F3')
    ws_stats.sheet_view.showGridLines = False

    # Auto-fit columns in all sheets
    autofit_columns_for_all_sheets(wb)

    wb.save(excel_path)
    wb.close()
    print('4.3) Excel export completed')
